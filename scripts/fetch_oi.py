"""
fetch_oi.py
JPX 当日取引高等 Excel から行使価格別OIを取得する
URL: https://www.jpx.co.jp/markets/derivatives/trading-volume/tvdivq00000014nn-att/{YYYYMMDD}_market_data_daytime.xlsx

Excelシート構造（market_data_OP）:
- 上部: ヘッダー（限月、指数値等）
- 左半分: CALL（5日平均比、今日、建玉残高前日比、前月比、横%、残高）
- 中央: 行使価格
- 右半分: PUT（同様）
"""

import os
import sys
import io
import json
import logging
import requests
import pandas as pd
import openpyxl
from datetime import date, timedelta
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)

# 2026/4/13以降: ファイル名が whole_day に変更
# 建玉残高専用ファイルは 20:00 公開（こちらを優先）
BASE_URL_OI         = "https://www.jpx.co.jp/markets/derivatives/trading-volume/tvdivq00000014nn-att/{date}open_interest.xlsx"
BASE_URL_WHOLE_DAY  = "https://www.jpx.co.jp/markets/derivatives/trading-volume/tvdivq00000014nn-att/{date}_market_data_whole_day.xlsx"
BASE_URL_OLD        = "https://www.jpx.co.jp/markets/derivatives/trading-volume/tvdivq00000014nn-att/{date}_market_data_daytime.xlsx"

# シート名（4/13以降: market_data_OP）
SHEET_OP_NEW   = "market_data_OP"
SHEET_OP_OLD   = "デリバティブ取引市況（日中）"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; JPX-OI-Bot/1.0)",
    "Referer": "https://www.jpx.co.jp/markets/derivatives/trading-volume/",
}


def prev_business_days(n=5):
    days, d = [], date.today()
    while len(days) < n:
        if d.weekday() < 5:
            days.append(d.strftime("%Y%m%d"))
        d -= timedelta(days=1)
    return days


def fetch_excel(target_date: str) -> tuple[bytes | None, str | None]:
    """
    建玉残高Excel を取得。優先順:
    1. open_interest.xlsx (建玉残高専用, 20:00公開)
    2. market_data_whole_day.xlsx (日通し, 17:30公開, 4/13以降)
    3. market_data_daytime.xlsx (旧形式)
    戻り値: (バイト列, 使用URL)
    """
    urls = [
        BASE_URL_OI.format(date=target_date),
        BASE_URL_WHOLE_DAY.format(date=target_date),
        BASE_URL_OLD.format(date=target_date),
    ]
    for url in urls:
        try:
            r = requests.get(url, headers=HEADERS, timeout=30)
            if r.status_code == 200:
                log.info("Excel取得成功: %s (%d bytes)", url, len(r.content))
                return r.content, url
            log.warning("HTTP %d: %s", r.status_code, url)
        except Exception as e:
            log.error("接続エラー: %s", e)
    return None, None


def find_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet | None:
    """オプション市況シートを探す"""
    for name in [SHEET_OP_NEW, SHEET_OP_OLD]:
        if name in wb.sheetnames:
            log.info("シート発見: %s", name)
            return wb[name]
    # 部分一致検索
    for name in wb.sheetnames:
        if "OP" in name or "オプション" in name or "option" in name.lower():
            log.info("シート発見(部分一致): %s", name)
            return wb[name]
    # open_interest.xlsx の場合、シート名が異なる可能性
    # 全シートを表示して部分一致で探す
    log.warning("オプションシートが見つかりません。利用可能シート: %s", wb.sheetnames)
    # フォールバック: 最初のシートを試す
    if wb.sheetnames:
        log.info("フォールバック: 最初のシート '%s' を使用", wb.sheetnames[0])
        return wb[wb.sheetnames[0]]
    return None


def parse_oi_sheet(ws) -> dict:
    """
    シートからOIデータを解析。
    行使価格列を探し、CALL OI・PUT OIを抽出。
    行使価格は数値で、CALL残高は行使価格の左側、PUT残高は右側にある。
    """
    all_values = [[cell.value for cell in row] for row in ws.iter_rows()]

    # 行使価格列を特定（大きな数値が縦に並ぶ列）
    strike_col = None
    strike_row_start = None

    for col_idx in range(len(all_values[0]) if all_values else 0):
        consecutive = 0
        first_row = None
        for row_idx, row in enumerate(all_values):
            val = row[col_idx] if col_idx < len(row) else None
            if isinstance(val, (int, float)) and 10000 <= val <= 100000:
                consecutive += 1
                if first_row is None:
                    first_row = row_idx
            else:
                if consecutive >= 5:
                    strike_col = col_idx
                    strike_row_start = first_row
                    break
                consecutive = 0
                first_row = None
        if strike_col is not None:
            break

    if strike_col is None:
        log.warning("行使価格列が見つかりません")
        return {}

    log.info("行使価格列: %d, 開始行: %d", strike_col, strike_row_start)

    # ヘッダー行を探して列構造を特定
    # CALL残高列: 行使価格列の左側にある「残高」ラベルの列
    # PUT残高列: 行使価格列の右側にある「残高」ラベルの列
    call_oi_col = None
    put_oi_col = None
    month_col = None

    # ヘッダー行（行使価格列の上の行）からラベルを探す
    for row_idx in range(max(0, strike_row_start - 5), strike_row_start):
        row = all_values[row_idx]
        for col_idx, val in enumerate(row):
            if val == "残高" or val == "建玉残高":
                if col_idx < strike_col and call_oi_col is None:
                    call_oi_col = col_idx
                elif col_idx > strike_col and put_oi_col is None:
                    put_oi_col = col_idx

    # ラベルで見つからない場合は位置から推定
    # 一般的にCALL残高は行使価格の2列左、PUT残高は2列右
    if call_oi_col is None:
        call_oi_col = strike_col - 1
    if put_oi_col is None:
        put_oi_col = strike_col + 1

    log.info("CALL OI列: %d, PUT OI列: %d", call_oi_col, put_oi_col)

    # 限月列を探す（OVER/UNDER行の前後に限月が書かれている）
    results = {}  # {month: {strike: {call_oi, put_oi}}}
    current_month = "202604"  # デフォルト

    for row_idx in range(strike_row_start, len(all_values)):
        row = all_values[row_idx]
        if not row or len(row) <= max(strike_col, put_oi_col):
            continue

        # 限月の更新（文字列で YYYYMM 形式または OVER/UNDER）
        for col_idx in range(min(strike_col, 3)):
            cell_val = row[col_idx] if col_idx < len(row) else None
            if isinstance(cell_val, str) and len(cell_val) == 6 and cell_val.isdigit():
                current_month = cell_val
                break

        strike_val = row[strike_col] if strike_col < len(row) else None
        if not isinstance(strike_val, (int, float)) or not (10000 <= strike_val <= 100000):
            continue

        call_oi = row[call_oi_col] if call_oi_col < len(row) else None
        put_oi  = row[put_oi_col]  if put_oi_col  < len(row) else None

        call_oi = float(call_oi) if isinstance(call_oi, (int, float)) else 0.0
        put_oi  = float(put_oi)  if isinstance(put_oi,  (int, float)) else 0.0

        if current_month not in results:
            results[current_month] = {}
        results[current_month][float(strike_val)] = {
            "call_oi": call_oi,
            "put_oi":  put_oi,
        }

    total_strikes = sum(len(v) for v in results.values())
    log.info("OI取得完了: %d限月, %d行使価格", len(results), total_strikes)
    return results


def calc_max_pain_from_oi(oi_data: dict, S: float) -> float:
    """実OIを使ったMaxPain計算（ATM±30%）"""
    if not oi_data:
        return S
    lower, upper = S * 0.70, S * 1.30
    strikes = {k: v for k, v in oi_data.items() if lower <= k <= upper}
    if not strikes:
        strikes = oi_data

    pain = {}
    for k_exp in strikes:
        total = 0.0
        for k, d in strikes.items():
            total += d["call_oi"] * max(0.0, k - k_exp)
            total += d["put_oi"]  * max(0.0, k_exp - k)
        pain[k_exp] = total

    return min(pain, key=pain.get) if pain else S


def calc_pcr_from_oi(oi_data: dict, S: float) -> float | None:
    """実OIを使ったPCR計算（ATM±30%）"""
    if not oi_data:
        return None
    lower, upper = S * 0.70, S * 1.30
    strikes = {k: v for k, v in oi_data.items() if lower <= k <= upper}
    if not strikes:
        return None

    call_sum = sum(d["call_oi"] for d in strikes.values())
    put_sum  = sum(d["put_oi"]  for d in strikes.values())

    if call_sum > 0:
        return round(put_sum / call_sum, 3)
    return None


def main():
    target_env = os.environ.get("TARGET_DATE", "").strip()
    candidates = [target_env] if target_env else prev_business_days(5)

    raw = None
    used_date = None
    used_url = None
    for d in candidates:
        raw, used_url = fetch_excel(d)
        if raw:
            used_date = d
            break

    if not raw:
        log.warning("Excel取得失敗 → OIデータなしでスキップ")
        sys.exit(0)

    log.info("使用ファイル: %s", used_url)

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = find_sheet(wb)
    if not ws:
        log.warning("シート解析失敗 → スキップ")
        sys.exit(0)

    oi_by_month = parse_oi_sheet(ws)

    # 先物価格（日経平均）をexisting JSONから取得
    meta_path = DATA_DIR / "meta.json"
    S = 50000.0
    if meta_path.exists():
        meta = json.loads(meta_path.read_text())
        S = meta.get("underlying_close", 50000.0)
    log.info("原資産価格: %.2f", S)

    # 月次限月のみ（6桁）
    monthly = {k: v for k, v in oi_by_month.items() if len(str(k).strip()) == 6}

    # 限月別MaxPain・PCR計算
    results = {}
    for month, oi_data in sorted(monthly.items()):
        mp  = calc_max_pain_from_oi(oi_data, S)
        pcr = calc_pcr_from_oi(oi_data, S)
        log.info("限月 %s: MaxPain=%.0f PCR=%s", month, mp, pcr)
        results[month] = {
            "max_pain": mp,
            "pcr":      pcr,
            "oi_by_strike": {
                str(k): {"call_oi": v["call_oi"], "put_oi": v["put_oi"]}
                for k, v in oi_data.items()
            },
        }

    # 保存
    out = {
        "generated_at": date.today().isoformat(),
        "excel_date":   used_date,
        "underlying_close": S,
        "months":       results,
    }
    out_path = DATA_DIR / "oi_latest.json"
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2))
    log.info("OI分析保存: %s", out_path)


if __name__ == "__main__":
    main()
